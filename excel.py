from openpyxl import load_workbook

wb = load_workbook('Excel_file_path.xlsx')   #the excel file you will save
ws = wb['Sayfa1']
yayın1 = "name of text"
yayın2 = "Second name or id nummber of text"
recall = ....    #you will update regarding to recall value
precision = .... #you will update regarding to precision value

#row,column values depend on your excel table, every 
ws.cell(row=102,column=3).value = yayın1                #name of text
ws.cell(row=102,column=4).value = yayın2                #Second name or id nummber of text
ws.cell(row=102,column=5).value = gercekmetinuzunlugu   #length of original text
ws.cell(row=102,column=6).value = özetuzunlugu          #length of summarization you produced
ws.cell(row=102,column=8).value = orjinalozet           #length of original abstract
ws.cell(row=102,column=9).value = recall                #recall value
ws.cell(row=102,column=10).value = precision            #precision value

wb.save('Excel_file_path.xlsx')



















            
