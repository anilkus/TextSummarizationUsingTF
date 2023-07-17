
import bs4 as bs
import urllib.request
import re
import nltk
from tkinter import *
import fitz
from tkinter import messagebox
import ssl 
from smtplib import SMTP



with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Abstracts/z_Anadolu Kliniği Tıp Bilimleri Dergisi/10.21673-anadoluklin.1218869-2831117.txt", 'r',encoding="utf-8") as text2:
    abstract=text2.read()
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/İşlenecek Kısım/z_Anadolu Kliniği Tıp Bilimleri Dergisi/10.21673-anadoluklin.1218869-2831117.txt", 'r',encoding="utf-8") as text:
    islenecek=text.read()
    article_text=islenecek
    #print(islenecek)
   
abstract=abstract.lower()
abstract = re.sub(r'\([^()]*\d+[^()]*\)', '', abstract)
abstract = re.sub(r'\[[^\[\]]*\d+[^\[\]]*\]', '', abstract)
#text = re.sub(r'\[[0-9]*\]', ' ', text)
#text = re.sub(r'\s+', ' ', text)  
article_text = article_text.lower()
#text = re.sub(r'\s+', ' ', text)
#text = re.sub(r'\([^)]*\)', '', text)
#text = re.sub(r'/',' ',text)
#text = re.sub("(\d+)","",text) 
article_text = re.sub(r'\([^()]*\d+[^()]*\)', '', article_text)
article_text = re.sub(r'\[[^\[\]]*\d+[^\[\]]*\]', '', article_text)
article_text = article_text.replace("ark.", '')
#paragraftaki cümle listesi , cümlelere ayırıyor
sentence_list = nltk.sent_tokenize(article_text)

#gereksiz kelimeler tespit eder ayırır
stopwords = nltk.corpus.stopwords.words('turkish')
#stopwords.append('ancak')
print(stopwords)
#print(article_text)
#keywords bulma

#print(keywords)

#sıklık
word_frequencies = {}
for word in nltk.word_tokenize(article_text):
    if word not in stopwords:
        if word not in word_frequencies.keys():
            word_frequencies[word] = 1
        else:
            word_frequencies[word] += 1
#word_frequencies['vertebra']*=100
#word_frequencies['anjiyotensin']*=100
#word_frequencies['psikolojik']*=100
#word_frequencies['olumsuz']*=100
#word_frequencies['ruminatif']*=100
#word_frequencies['ders']*=100
#word_frequencies['koronavirüs']*=100


#kelime ağırlını hesaplar            
#print(word_frequencies)
maximum_frequncy = max(word_frequencies.values())
#print(maximum_frequncy)
for word in word_frequencies.keys():
    word_frequencies[word] = (word_frequencies[word]/maximum_frequncy)


#print(word_frequencies)
#cümle ağırlığını hesaplar

sentence_scores = {}
for sent in sentence_list:
    for word in nltk.word_tokenize(sent.lower()):
        if word in word_frequencies.keys():
            #if len(sent.split(' ')) < 30:
                if sent not in sentence_scores.keys():
                    sentence_scores[sent] = word_frequencies[word]
                else:
                    sentence_scores[sent] += word_frequencies[word]

#print(sentence_scores)                    
import heapq
summary_sentences = heapq.nlargest(15, sentence_scores, key=sentence_scores.get)
#print("Gerçek Metin:")
#print(article_text)
summary = ' '.join(summary_sentences)
print("Özetlenmiş Metin:")
print(summary)
print("Gerçek Metin Uzunluğu (Karakter): {}".format(len(article_text)))
gercekmetinuzunlugu = len(article_text)

print("Özet Uzunluğu (Karakter): {}".format(len(summary)))                    
özetuzunlugu = len(summary)
print(len(abstract))
orjinalozet=len(abstract)
from rouge import Rouge
ROUGE = Rouge()

print(ROUGE.get_scores(summary, abstract))
#print(abstract)

'''
ozet = summary
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Çalışmanın Özetleri/TF/z_Anadolu Kliniği Tıp Bilimleri Dergisi/10.21673-anadoluklin.1218869-2831117.txt", "w",encoding="utf-8") as file:
    file.write(ozet)
'''

from openpyxl import load_workbook
'''
wb = load_workbook('C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Sonuçlar/TF.xlsx')
ws = wb['Sayfa1']
yayın1 = "anadolu yeni"
yayın2 = "2"
recall = 0.48
precision = 0.26
ws.cell(row=102,column=3).value = yayın1
ws.cell(row=102,column=4).value = yayın2
ws.cell(row=102,column=5).value = gercekmetinuzunlugu
ws.cell(row=102,column=6).value = özetuzunlugu
ws.cell(row=102,column=8).value = orjinalozet
ws.cell(row=102,column=9).value = recall
ws.cell(row=102,column=10).value = precision
wb.save('C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Sonuçlar/TF.xlsx')

'''


















            